# import openpyxl and tkinter modules
from openpyxl import *
from tkinter import *
from tkinter import ttk

# globally declare wb and sheet variable

# opening the existing excel file
wb = load_workbook('C:\\Users\\PC\\Documents\\Open.xlsx')

# create the sheet object
sheet = wb.active


def excel():
	
	# resize the width of columns in
	# excel spreadsheet
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 40
    sheet.column_dimensions['G'].width = 30
    sheet.column_dimensions['H'].width = 20

	# write given data to an excel spreadsheet
	# at particular location
    sheet.cell(row=1, column=1).value = "Mã số sinh viên"
    sheet.cell(row=1, column=2).value = "Họ tên"
    sheet.cell(row=1, column=3).value = "Ngày sinh"
    sheet.cell(row=1, column=4).value = "Email"
    sheet.cell(row=1, column=5).value = "Số điện thoại"
    sheet.cell(row=1, column=6).value = "Học kỳ"
    sheet.cell(row=1, column=7).value = "Năm học"
    sheet.cell(row=1, column=8).value = "Chọn môn học"
    
def focus1(event):
	mssv_field.focus_set()
def focus2(event):
    ngaySinh_field.focus_set()
def focus3(event):
    email_field.focus_set()
def focus4(event):
    sdt_field.focus_set()
def focus5(event):
    hk_field.focus_set()
def focus6(event):
   namHoc_field.focus_set()
def focus7(event):
    chonMonHoc_field.focus_set()
 
def clear():
    mssv_field.delete(0,END)
    ten_field.delete(0,END)
    ngaySinh_field.delete(0,END)
    email_field.delete(0,END)
    sdt_field.delete(0,END)
    hk_field.delete(0,END)
    namHoc_field.delete(0,END)
    chonMonHoc_field.delete(0,END)
def insert():
    if (mssv_field.get() == "" and
		ten_field.get() == "" and
        ngaySinh_field.get() == "" and
        email_field.get() == "" and
        sdt_field.get() == "" and
        hk_field.get() == "" and
        namHoc_field.get() == ""and
        chonMonHoc_field.get()):
        print("empty input")
    else:

		# assigning the max row and max column
		# value upto which data is written
		# in an excel sheet to the variable
        current_row = sheet.max_row
        current_column = sheet.max_column

		# get method returns current text
		# as string which we write into
		# excel spreadsheet at particular location
        sheet.cell(row=current_row + 1, column=1).value = mssv_field.get()
        sheet.cell(row=current_row + 1, column=2).value = ten_field.get()
        sheet.cell(row=current_row + 1, column=3).value = ngaySinh_field.get()
        sheet.cell(row=current_row + 1, column=4).value = email_field.get()
        sheet.cell(row=current_row + 1, column=5).value = sdt_field.get()
        sheet.cell(row=current_row + 1, column=6).value = hk_field.get()
        sheet.cell(row=current_row + 1, column=7).value = namHoc_field.get()
        sheet.cell(row=current_row + 1, column=8).value = chonMonHoc_field.get()

		# save the file
        wb.save('C:\\Users\\PC\\Documents\\Open.xlsx')

		# set focus on the name_field box
        ten_field.focus_set()

		# call the clear() function
        clear()
if __name__ == "__main__":
	
	# create a GUI window
    root = Tk()

	# set the background colour of GUI window
    root.configure(background='light green')

	# set the title of GUI window
    root.title("registration form")

	# set the configuration of GUI window
    root.geometry("500x300")

    excel()
    # create a Form label
    heading = Label(root, text="Form", bg="light green")

	# create a Name label
    mssv = Label(root, text="Mã số sinh viên", bg="light green")

	# create a Course label
    ten = Label(root, text="Họ và tên", bg="light green")

	# create a Semester label
    ngaySinh = Label(root, text="Ngày sinh", bg="light green")

	# create a Form No. label
    email = Label(root, text="Email", bg="light green")

	# create a Contact No. label
    sdt = Label(root, text="Số điện thoại", bg="light green")

	# create a Email id label
    hk = Label(root, text="Học kỳ", bg="light green")

	# create a address label
    namHoc = Label(root, text="Năm học", bg="light green")
    
    chonMonHoc = Label(root,text = "Chọn môn học", bg="light green")
    
    heading.grid(row=0, column=1)
    mssv.grid(row=1, column=0)
    ten.grid(row=2, column=0)
    ngaySinh.grid(row=3, column=0)
    email.grid(row=4, column=0)
    sdt.grid(row=5, column=0)
    hk.grid(row=6, column=0)
    namHoc.grid(row=7, column=0)
    chonMonHoc.grid(row=8,column = 0)

    mssv_field = Entry(root)
    ten_field = Entry(root)
    ngaySinh_field = Entry(root)
    email_field = Entry(root)
    sdt_field = Entry(root)
    hk_field = Entry(root)
    namHoc_field = ttk.Combobox(root,width=17)
    chonMonHoc_field = Label(root)
    var1 = IntVar()
    Checkbutton(root,text="Lập trình Python", variable=1, background="light green").place(x=90, y=175)
    var2 = IntVar()
    Checkbutton(root,text="Lập trình Java", variable=2, background="light green").place(x=250, y=175)
    var3 = IntVar()
    Checkbutton(root,text="Công nghệ phần mềm", variable=3, background="light green").place(x=90, y=220)
    var4 = IntVar()
    Checkbutton(root,text="Phát triển ứng dụng", variable=4, background="light green").place(x=250, y=220)
    
    
    namHoc_field['values'] = ('2022 - 2023',
                              '2023 - 2024')

    mssv_field.bind("<Return>", focus1)
    ten_field.bind("<Return>", focus2)
    ngaySinh_field.bind("<Return>", focus1)
    email_field.bind("<Return>", focus1)
    sdt_field.bind("<Return>", focus1)
    hk_field.bind("<Return>", focus1)
    namHoc_field.bind("<Return>", focus1)
    chonMonHoc_field.bind("<Return>", focus1)
    
    mssv_field.grid(row=1, column=1, ipadx="100")
    ten_field.grid(row=2, column=1, ipadx="100")
    ngaySinh_field.grid(row=3, column=1, ipadx="100")
    email_field.grid(row=4, column=1, ipadx="100")
    sdt_field.grid(row=5, column=1, ipadx="100")
    hk_field.grid(row=6, column=1, ipadx="100")
    namHoc_field.grid(row=7, column=1, ipadx="100")

    excel()
 
    # create a Submit Button and place into the root window
    resign = Button(root, text="Đăng ký", fg="Black",
                            bg="Green", command=insert)
    resign.place(x=125, y=270)
    
    quit = Button(root, text="Thoát", fg="Black",
                            bg="Green", command=quit)
    quit.place(x=300, y=270)
 
 
    # start the GUI
    root.mainloop()